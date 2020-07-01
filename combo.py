from textblob import TextBlob
import pandas as pd
import openpyxl

def sentiment(fill):
    fi = fill
    print(fi)
    df = pd.read_excel(fi)
    df = df.filter(items=['Unnamed: 1'])
    df = df.astype(str)
    text = df.values.tolist()
    j = 2
    wb = openpyxl.load_workbook(fi)
    ws = wb['Sheet1']
    column = ws.max_column + 1
    for i in text:
        obj = TextBlob("".join(i))
        sentiment = obj.sentiment.polarity
        if sentiment == 0:
            print('review is neutral')
            a = 'review is neutral'
        elif sentiment > 0:
            print('review is positive')
            a = 'review is positive'
        else:
            print('review is negative')
            a = 'review is negative'

        ws.cell(row=j, column=column, value=sentiment)
        ws.cell(row=j, column=column + 1, value=a)
        j = j + 1
    wb.save(fi)
    wb.close()
    from output import sen
    sen(fi)
    print("combo file ")
    return None
