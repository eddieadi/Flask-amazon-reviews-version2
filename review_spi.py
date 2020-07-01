import scrapy
from openpyxl import Workbook
from scrapy.crawler import CrawlerProcess
from combo import sentiment
import random

file = 'report_'+str(random.randint(1,200))+'.xlsx'
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

def rr():
    return file

class ReviewItem(scrapy.Item):
    name = scrapy.Field()
    inshort = scrapy.Field()
    rating = scrapy.Field()
    pass

class ReviewSpiSpider(scrapy.Spider):
    name = 'revieww'
    file1 = open("url.txt", "r")
    url = file1.read()
    file1.close()
    print("spider inside   "+url)
    start_urls = [url]

    def parse(self, response):
        items = ReviewItem()

        name = response.css('#cm_cr-review_list .a-profile-name').css('::text').extract()
        inshort = response.css('.a-text-bold span').css('::text').extract()
        rating = response.css('#cm_cr-review_list .review-rating').css('::text').extract()

        items['name'] = name
        items['inshort'] = inshort
        items['rating'] = rating
        yield items
        print(type(name))

        row = ws.max_row
        r2 = ws.max_row
        r3 = ws.max_row

        for i in name:
            column = 1
            row += 1
            ws.cell(row = row, column = column, value = i)

        for j in inshort:
            column = 2
            r2 += 1
            ws.cell(row=r2, column=column, value=j)

        for k in rating:
            column = 3
            r3 += 1
            k = k[0:3:]
            ws.cell(row=r3, column=column, value=k)

        nextp = response.css('li.a-last a::attr(href)').get()
        if nextp is not None:
            yield response.follow(nextp, callback=self.parse)
        else:
            print("over the end  " +file)
            wb.save(file)
            sentiment(file)
            print("spider file")
            return None

if __name__ == '__main__':
    # run scraper

    process = CrawlerProcess()
    process.crawl(ReviewSpiSpider)
    process.start()